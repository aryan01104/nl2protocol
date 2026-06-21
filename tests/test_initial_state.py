"""Tests for the initial-state spreadsheet parser (one clause of the
parse_initial_state contract per test)."""

from io import BytesIO

import openpyxl
import pytest

from nl2protocol.stage_1_pre_extraction.initial_state import (
    InitialStateSheet,
    parse_initial_state,
)

CONFIG = {"labware": {"source_plate": {"load_name": "x"},
                      "reagent_reservoir": {"load_name": "y"}}}


def _wb_bytes(blocks):
    """blocks: list of (anchor_row, anchor_col, header, colnums, rows) — 1-based
    anchor; rows is [(letter, [values aligned to colnums])]."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for ar, ac, header, colnums, rowdata in blocks:
        ws.cell(row=ar, column=ac, value=header)
        for j, cn in enumerate(colnums):
            ws.cell(row=ar + 1, column=ac + 1 + j, value=cn)
        for i, (letter, vals) in enumerate(rowdata):
            ws.cell(row=ar + 2 + i, column=ac, value=letter)
            for j, v in enumerate(vals):
                if v is not None:
                    ws.cell(row=ar + 2 + i, column=ac + 1 + j, value=v)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _valid_stacked():
    return _wb_bytes([
        (1, 1, "source_plate — substance", [1, 2],
         [("A", ["dye", "dye"]), ("B", ["lysis", "lysis"])]),
        (6, 1, "source_plate — volume (uL)", [1, 2],
         [("A", [10, 10]), ("B", [50, 50])]),
    ])


def test_valid_sheet_pairs_substance_and_volume():
    sheet = parse_initial_state(_valid_stacked(), CONFIG)
    assert sheet.errors == []
    assert sheet.cells[("source_plate", "A1")] == ("dye", 10.0)
    assert sheet.cells[("source_plate", "B2")] == ("lysis", 50.0)
    assert len(sheet.cells) == 4


def test_volume_is_coerced_to_float():
    sheet = parse_initial_state(_valid_stacked(), CONFIG)
    _sub, vol = sheet.cells[("source_plate", "A1")]
    assert isinstance(vol, float)


def test_side_by_side_blocks_parse():
    data = _wb_bytes([
        (1, 1, "source_plate — substance", [1], [("A", ["dye"])]),
        (1, 5, "source_plate — volume", [1], [("A", [10])]),
    ])
    sheet = parse_initial_state(data, CONFIG)
    assert sheet.errors == []
    assert sheet.cells == {("source_plate", "A1"): ("dye", 10.0)}


def test_unknown_label_is_reported_not_raised():
    data = _wb_bytes([
        (1, 1, "mystery_plate — substance", [1], [("A", ["dye"])]),
        (6, 1, "mystery_plate — volume", [1], [("A", [10])]),
    ])
    sheet = parse_initial_state(data, CONFIG)
    assert sheet.cells == {}
    assert any("mystery_plate" in e and "not a labware" in e for e in sheet.errors)


def test_non_numeric_volume_reported_others_kept():
    data = _wb_bytes([
        (1, 1, "source_plate — substance", [1, 2], [("A", ["dye", "lysis"])]),
        (6, 1, "source_plate — volume", [1, 2], [("A", ["lots", 50])]),
    ])
    sheet = parse_initial_state(data, CONFIG)
    assert ("source_plate", "A2") in sheet.cells
    assert ("source_plate", "A1") not in sheet.cells
    assert any("not a number" in e for e in sheet.errors)


def test_well_in_one_map_only_reported():
    data = _wb_bytes([
        (1, 1, "source_plate — substance", [1, 2], [("A", ["dye", "lysis"])]),
        (6, 1, "source_plate — volume", [1, 2], [("A", [10, None])]),
    ])
    sheet = parse_initial_state(data, CONFIG)
    assert ("source_plate", "A1") in sheet.cells
    assert ("source_plate", "A2") not in sheet.cells
    assert any("A2" in e and "no volume" in e for e in sheet.errors)


def test_duplicate_block_reported():
    data = _wb_bytes([
        (1, 1, "source_plate — substance", [1], [("A", ["dye"])]),
        (6, 1, "source_plate — substance", [1], [("A", ["other"])]),
        (11, 1, "source_plate — volume", [1], [("A", [10])]),
    ])
    sheet = parse_initial_state(data, CONFIG)
    assert any("duplicate substance" in e for e in sheet.errors)


def test_unreadable_bytes_yield_error_not_exception():
    sheet = parse_initial_state(b"not a workbook", CONFIG)
    assert sheet.cells == {}
    assert len(sheet.errors) == 1
