#!/usr/bin/env python3
"""
evals/diff_state.py — see how each step changes across pipeline stages, in Excel.

The pipeline dumps its full state to `output/pipeline_state_<ts>.json`, keyed by
stage (`stage_2_extraction`, `stage_3_gap_resolver`, `stage_5_spec`, ...). The
extraction stage and the final spec both carry a `steps` list with identical
fields, so the *effect* of gap-filling/resolution is exactly the diff between
them.

This tool aligns two stages' steps by their `order` field, flattens each step's
nested dicts (volume/substance/source/...) into dotted paths, and writes a
grouped .xlsx:

    | Step (merged) | Field        | <from>      | <to>        |
    | step 1        | action       | transfer    | transfer    |
    |               | volume.value | 80.0        | 80.0        |
    |               | source.well  | A1          | A1          |

The leftmost cell is merged vertically over all of a step's field rows. Rows
whose value changed between the two stages are highlighted yellow, so the
gap-filling pops out. Header row is bold and frozen.

Usage:
    python evals/diff_state.py                      # newest dump, extraction -> spec
    python evals/diff_state.py path/to/state.json   # a specific dump
    python evals/diff_state.py --from stage_2_extraction --to stage_5_spec
    python evals/diff_state.py --out report.xlsx    # choose output path
    python evals/diff_state.py --list               # list stages in the dump

Requires openpyxl (`venv/bin/python -m pip install openpyxl`).
"""

import argparse
import glob
import json
import os
import sys

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")


def newest_state_file():
    """Return the most recently modified pipeline_state_*.json, or None."""
    matches = glob.glob(os.path.join(OUTPUT_DIR, "pipeline_state_*.json"))
    return max(matches, key=os.path.getmtime) if matches else None


def flatten(node, prefix=""):
    """Flatten nested dicts to {dotted.path: value}.

    Dicts recurse so leaf paths like `volume.value` and `source.well` each get
    their own row. Lists and scalars are left as leaves (rendered at write time)
    rather than exploded, to keep the table readable.
    """
    out = {}
    if isinstance(node, dict):
        for k, v in node.items():
            path = f"{prefix}.{k}" if prefix else str(k)
            if isinstance(v, dict):
                out.update(flatten(v, path))
            else:
                out[path] = v
    else:
        out[prefix] = node
    return out


def by_order(steps):
    """Map each step's `order` to the step dict; fall back to list position."""
    return {(s.get("order", i) if isinstance(s, dict) else i): s
            for i, s in enumerate(steps)}


def cell_value(v):
    """Render a leaf for an Excel cell: None -> '', list/dict -> JSON, else as-is."""
    if v is None:
        return ""
    if isinstance(v, (list, dict)):
        return json.dumps(v)
    return v


def default_out_path(state_path, from_stage, to_stage):
    """Derive a sibling .xlsx path from the input filename."""
    base = os.path.basename(state_path)
    stem = base[:-5] if base.endswith(".json") else base  # drop .json
    stem = stem.replace("pipeline_state", "step_diff")
    return os.path.join(os.path.dirname(state_path), f"{stem}.xlsx")


def build_workbook(state, from_stage, to_stage):
    """Build the grouped diff workbook. Returns (Workbook, changed_rows, total_rows)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    before = by_order(state[from_stage]["steps"])
    after = by_order(state[to_stage]["steps"])

    wb = Workbook()
    ws = wb.active
    ws.title = "diff"[:31]

    # Columns: A=Step  B=Field (path level 1)  C=Subfield (level 2)
    #          D=Sub-subfield (level 3+)  E=from  F=to. The Step/Field/Subfield
    #          cells are merged vertically over the rows that share their prefix,
    #          so the sheet reads like an outline instead of dotted paths.
    ws.append(["Step", "Field", "Subfield", "Sub-subfield", from_stage, to_stage])
    hdr_fill = PatternFill("solid", fgColor="305496")
    hdr_font = Font(bold=True, color="FFFFFF")
    for c in ws[1]:
        c.fill, c.font = hdr_fill, hdr_font
        c.alignment = Alignment(vertical="center")

    changed_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    top_align = Alignment(vertical="top", wrap_text=True)
    grp_align = Alignment(vertical="center", horizontal="left")
    NCOLS = 6

    def split3(path):
        """Split a dotted path into (level1, level2, level3+) cells.

        Levels beyond 3 collapse into the third cell so the table never grows
        wider than the three grouping columns. Missing levels are ''.
        """
        segs = path.split(".")
        return (segs[0],
                segs[1] if len(segs) > 1 else "",
                ".".join(segs[2:]) if len(segs) > 2 else "")

    def merge_label(col, r0, r1, text, font, align):
        """Merge a grouping column over rows r0..r1 and label its top cell."""
        if r1 > r0:
            ws.merge_cells(start_row=r0, start_column=col, end_row=r1, end_column=col)
        c = ws.cell(r0, col, text)
        c.font = font
        c.alignment = align

    def merge_runs(col, row_keys, label_of, font):
        """Merge consecutive rows sharing a key; label each run via label_of(key).

        Runs whose label is '' (e.g. a bare top-level field with no subfield)
        are left as single unmerged cells.
        """
        i, n = 0, len(row_keys)
        while i < n:
            j = i
            key = row_keys[i][1]
            while j + 1 < n and row_keys[j + 1][1] == key:
                j += 1
            label = label_of(key)
            if label != "":
                merge_label(col, row_keys[i][0], row_keys[j][0], label, font, grp_align)
            i = j + 1

    changed_rows = 0
    row = 2
    orders = sorted(set(before) | set(after), key=lambda x: (isinstance(x, str), x))
    for order in orders:
        b = flatten(before.get(order, {}))
        a = flatten(after.get(order, {}))
        step_start = row
        paths = sorted(set(b) | set(a))

        l1_keys, l2_keys = [], []   # (row, key) for Field and Subfield merges
        for path in paths:
            c1, c2, c3 = split3(path)
            bv, av = b.get(path), a.get(path)
            ws.cell(row, 4, c3)
            ws.cell(row, 5, cell_value(bv))
            ws.cell(row, 6, cell_value(av))
            changed = bv != av
            if changed:
                changed_rows += 1
            for col in range(1, NCOLS + 1):
                c = ws.cell(row, col)
                c.border = border
                if col >= 4:
                    c.alignment = top_align
                if changed:
                    c.fill = changed_fill
            l1_keys.append((row, c1))
            l2_keys.append((row, (c1, c2)))   # include c1 so runs reset per field
            row += 1

        # Merge Field (level 1) and Subfield (level 2) over their shared prefixes.
        merge_runs(2, l1_keys, lambda k: k, Font(bold=False))
        merge_runs(3, l2_keys, lambda k: k[1], Font(bold=False))
        # Merge the Step column over this step's whole block.
        merge_label(1, step_start, row - 1, f"step {order}", Font(bold=True),
                    Alignment(vertical="center", horizontal="center"))

    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 44
    ws.column_dimensions["F"].width = 44
    ws.freeze_panes = "A2"
    return wb, changed_rows, row - 2


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("state_file", nargs="?",
                   help="pipeline_state JSON (default: newest in output/)")
    p.add_argument("--from", dest="from_stage", default="stage_2_extraction",
                   help="stage to treat as 'before' (default: stage_2_extraction)")
    p.add_argument("--to", dest="to_stage", default="stage_5_spec",
                   help="stage to treat as 'after' (default: stage_5_spec)")
    p.add_argument("--out", help="output .xlsx path (default: sibling of input)")
    p.add_argument("--list", action="store_true",
                   help="list the stages present in the dump and exit")
    args = p.parse_args()

    path = args.state_file or newest_state_file()
    if not path:
        print("No pipeline_state_*.json found in output/. Run the pipeline first.",
              file=sys.stderr)
        return 1
    if not os.path.exists(path):
        print(f"No such file: {path}", file=sys.stderr)
        return 1

    with open(path) as f:
        state = json.load(f)
    print(f"state: {os.path.relpath(path, PROJECT_ROOT)}")

    if args.list:
        print("stages:")
        for k, v in state.items():
            n = len(v["steps"]) if isinstance(v, dict) and "steps" in v else ""
            print(f"  {k}{f'   (steps: {n})' if n != '' else ''}")
        return 0

    for stage in (args.from_stage, args.to_stage):
        v = state.get(stage)
        if not isinstance(v, dict) or "steps" not in v:
            print(f"'{stage}' has no steps list. Try --list.", file=sys.stderr)
            return 1

    try:
        wb, changed, total = build_workbook(state, args.from_stage, args.to_stage)
    except ImportError:
        print("openpyxl not installed. Run: venv/bin/python -m pip install openpyxl",
              file=sys.stderr)
        return 1

    out = args.out or default_out_path(path, args.from_stage, args.to_stage)
    wb.save(out)
    print(f"wrote {os.path.relpath(out, PROJECT_ROOT)}  "
          f"({changed}/{total} field-rows changed, {args.from_stage} -> {args.to_stage})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
